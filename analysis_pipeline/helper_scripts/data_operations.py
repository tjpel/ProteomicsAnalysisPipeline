import pandas as pd
import numpy as np
import json
import os
from collections.abc import Callable
import scipy.stats as stats
from statsmodels.stats.multitest import multipletests

from openpyxl.styles import PatternFill, Alignment

from .debug_wrapper import debug

def get_config():
    with open('analysis_pipeline/config.json') as config_file:
        config = json.load(config_file)

    return config

config = get_config()

def get_filetype_info(file_type: str) -> int:
    match file_type:
        case "Olink":
            return 5, "Assay", (1, 5)
        case "TMT Phospho":
            return 6, "Modifications in Master Proteins", (1, 0)
        
DIR_PATH = config['project_information']['relative_path']
METADATA_COL, PRIMARY_KEY_COL, SHEET1_FREEZE_PANES = get_filetype_info(config['project_information']['file_type'])

def load_intermediate_csv(prot_intermediate_id: str):
    if prot_intermediate_id == "sample_info":
        return pd.read_csv(os.path.join(DIR_PATH, config["paths"][prot_intermediate_id]))
    
    return pd.read_csv(os.path.join(DIR_PATH, config["paths"][prot_intermediate_id]), index_col=PRIMARY_KEY_COL)

@debug
def prepare_group_filtered_dict(prot_intermediate_id: str, study_groups: dict) -> dict:

    protein_count_data = load_intermediate_csv(prot_intermediate_id)
    sample_data = pd.read_csv(os.path.join(DIR_PATH, config["paths"]['sample_info']))

    prot_with_sample_info = get_prot_with_sample_info(
        protein_count_data,
        sample_data,
    )

    group_filtered_datasets = {}

    for group_num in study_groups.keys():
        if group_num == "Name":
            pass

        filter_items = study_groups[group_num] if isinstance(study_groups[group_num], dict) else None

        filtered_df = prot_with_sample_info.copy()
        if filter_items:
            for study_group_id, study_group in filter_items.items():
                filtered_df = filtered_df[filtered_df[study_group_id].isin(study_group.split(','))]

        group_filtered_datasets[group_num] = {}
        
        filtered_df.insert(3, "Group Number", group_num)

        if filter_items:
            group_filtered_datasets[group_num]["Dataset"] = filtered_df.T
            group_filtered_datasets[group_num]["Dataset"].columns = filtered_df.T.iloc[0, :]
            group_filtered_datasets[group_num]["Dataset"] = group_filtered_datasets[group_num]["Dataset"].iloc[1:, :]

            group_filtered_datasets[group_num]["Filters"] = filter_items

    return group_filtered_datasets

def sort_key(col):
    if "Comparison" in col:
        # Extract comparison number
        comp_num = int(col.split(":")[0].split()[-1])
        # Determine type priority: Mean < Log2 Fold Change < P-value
        if "Mean" in col:
            col_type = 0
        elif "Log2 Fold Change" in col:
            col_type = 1
        elif "P-value" in col and (not "FDR" in col):
            col_type = 2
        elif "P-value" in col:
            col_type = 3
        elif "Missing Values" in col:
            col_type = 4
        else:
            col_type = 99  # Default fallback if type is unrecognized
        return (0, comp_num, col_type, col)
    else:
        return (1, col)

@debug
def get_data_from_raw(raw_data_path: str):
    protein_and_meta_data = pd.read_excel(raw_data_path, 0)

    protein_meta_data = protein_and_meta_data.iloc[:, :METADATA_COL].copy(deep=True)
    protein_count_data = protein_and_meta_data.iloc[:, METADATA_COL:].copy(deep=True)
    protein_count_data.set_index(protein_meta_data[PRIMARY_KEY_COL], inplace=True)

    sample_info = pd.read_excel(raw_data_path, 1)
    study_group_info =  pd.read_excel(raw_data_path, 2)

    return protein_count_data, protein_meta_data, sample_info, study_group_info

@debug
def get_prot_with_sample_info(protein_count_data: pd.DataFrame, sample_info: pd.DataFrame) -> pd.DataFrame:

    protein_count_data_T = protein_count_data.T
    protein_count_data_T.reset_index(inplace=True)
    protein_count_data_T.rename(columns={'index':'Sample ID'}, inplace=True)

    prot_with_sample_info = pd.merge(sample_info, protein_count_data_T, on=sample_info.columns[0], how='outer')

    return prot_with_sample_info

@debug
def prepare_analysis_dataset() -> pd.DataFrame:

    #If we have log transformed the data previously, we'll then use the transformed data to calculate the following statistics
    #If we haven't previously transformed the data, we'll use the non-transformed data to calculate the following statistics

    output_df = load_intermediate_csv('post_pipeline')

    if os.path.exists(os.path.join(DIR_PATH, config['paths']['pre_transformation'])):
        alr_transformed = True
    else:
        alr_transformed = False

    for comparison_num, study_groups in config['comparisons'].items():

        group_filtered_datasets = prepare_group_filtered_dict('post_pipeline', study_groups)

        groups = [group for group in group_filtered_datasets.keys() if group != "Name"]

        for i in range(len(groups)-1):
            group1_name = groups[i]
            group1 = group_filtered_datasets[group1_name]['Dataset']
            for j in range(i + 1, len(groups)):
                group2_name = groups[j]
                group2 = group_filtered_datasets[group2_name]['Dataset']
                
                t_test_p_values = {}
                group1_means = {}
                group2_means = {}
                group1_missing = {}
                group2_missing = {}
                for protein in list(group1.index):
                    if isinstance(group1.loc[protein], pd.DataFrame) or isinstance(group2.loc[protein], pd.DataFrame):
                        print(f"Values in primary key column ({PRIMARY_KEY_COL}) are not unique. Please remove duplicate values in this column. This may be done in the configuration file by adding the 'Drop Duplicates' pipeline step.")
                        print(f"Duplicate values: {group1[group1.index.duplicated()].index} and/or {group2[group2.index.duplicated()].index}")
                        quit()
                    
                    group1_protein = pd.to_numeric(group1.loc[protein], errors='coerce')
                    group2_protein = pd.to_numeric(group2.loc[protein], errors='coerce')

                    t_test_p_values[protein] = stats.ttest_ind(group1_protein, group2_protein, nan_policy='omit').pvalue

                    """
                    if config["analysis_behavior"]["mean_type"] == 'Geometric':
                        try:
                            group1_means[protein] = stats.gmean((group1_protein), nan_policy='omit')
                            group2_means[protein] = stats.gmean((group2_protein), nan_policy='omit')
                        except Exception as e:
                            print(f"Error in protein {protein}: {e}")
                    else:
                    """
                    group1_means[protein] = group1_protein.mean()
                    group2_means[protein] = group2_protein.mean()

                    group1_missing[protein] = f"{group1_protein.isna().sum()} / {len(group1_protein)}"
                    group2_missing[protein] = f"{group2_protein.isna().sum()} / {len(group2_protein)}"
                
                p_values_series = pd.Series(t_test_p_values)
                valid_mask = p_values_series.notna()

                # Apply FDR correction only on valid (non-NaN) p-values
                _, p_vals_corrected, _, _ = multipletests(p_values_series[valid_mask].values, alpha=0.05, method='fdr_bh')

                # Create a new series with NaNs in the correct positions
                corrected_p_values_series = pd.Series(np.nan, index=p_values_series.index)
                corrected_p_values_series[valid_mask] = p_vals_corrected

                output_df[f"Comparison {comparison_num}: {group1_name} Mean"] = pd.Series(group1_means)
                output_df[f"Comparison {comparison_num}: {group2_name} Mean"] = pd.Series(group2_means)
                output_df[f"Comparison {comparison_num}: {group1_name} v. {group2_name} P-value"] = pd.Series(t_test_p_values)
                output_df[f"Comparison {comparison_num}: {group1_name} v. {group2_name} FDR-adj. P-value"] = pd.Series(corrected_p_values_series, index=p_values_series.index)
                output_df[f"Comparison {comparison_num}: {group1_name} Missing Values"] = pd.Series(group1_missing)
                output_df[f"Comparison {comparison_num}: {group2_name} Missing Values"] = pd.Series(group2_missing)

                if alr_transformed or config['project_information']['file_type'] == "Olink":
                    output_df[f"Comparison {comparison_num}: {group1_name} v. {group2_name} Log2 Fold Change"] = pd.Series(group1_means) - pd.Series(group2_means)
                else:
                    output_df[f"Comparison {comparison_num}: {group1_name} v. {group2_name} Log2 Fold Change"] = np.log2(pd.Series(group1_means) / pd.Series(group2_means))

    output_df = output_df[sorted(output_df.columns, key=sort_key)]

    output_df.to_csv(os.path.join(DIR_PATH, config['paths']['fold_change_dataset']))

    return output_df

#This method removes proteins that are missing more than the threshold amount in every (i.e. all) study group of each type of study group
@debug
def remove_missing_proteins_groupwise(protein_count_data: pd.DataFrame, sample_info: pd.DataFrame, threshold: float = 50) -> pd.DataFrame:
    threshold = threshold / 100 #change % to decimal

    # Prepare the protein count data to be merged with the sample info data
    protein_count_data_T = protein_count_data.T
    protein_count_data_T.reset_index(inplace=True)
    protein_count_data_T.rename(columns={'index': 'Sample ID'}, inplace=True)

    # Merge protein counts and sample info
    prot_with_sample_info = pd.merge(sample_info, protein_count_data_T, on=sample_info.columns[0], how='outer')

    # Fill NAs and NaNs with 0s
    prot_with_sample_info.fillna(0, inplace=True)

    # DataFrame to store dropped proteins
    dropped_proteins = pd.DataFrame(columns=[PRIMARY_KEY_COL])

    # Search through each study group for proteins that are largely missing for that group
    for study_group_col in sample_info.columns[1:]:
        # Create a set to track the proteins we want to remove
        indices_with_missing_total = {}

        # Latch to create the list if it doesn't exist (i.e., the first pass of this loop)
        establish_list = True

        for study_group_val in sample_info[study_group_col].unique():
            # Filter the dataset to look at just the study group of interest
            filtered_df = prot_with_sample_info[prot_with_sample_info[study_group_col] == study_group_val]

            # Find the ratio of how many proteins are missing in each column of the filtered dataset
            missing_ratio = filtered_df.eq(0).mean(axis=0)

            if establish_list:
                # Create a set of all proteins missing more than the threshold for the first study group we look at
                indices_with_missing_total = set(missing_ratio[missing_ratio > threshold].index)
                establish_list = False
            else:
                # Remove proteins from the set that are present in the following groups more than the threshold
                indices_with_missing_total -= set(missing_ratio[missing_ratio <= threshold].index)

        # Log dropped proteins for this study group type
        dropped = missing_ratio[missing_ratio > threshold].index
        for prot in dropped:
            dropped_proteins = pd.concat([
                dropped_proteins,
                pd.DataFrame({PRIMARY_KEY_COL: [prot]})
            ], ignore_index=True)

        # Drop proteins missing in all study groups of that type
        prot_with_sample_info.drop(columns=indices_with_missing_total, inplace=True)

    # Save dropped proteins to CSV
    dropped_path = os.path.join(config["project_information"]["relative_path"], config["paths"]["dropped_proteins_list"])
    dropped_proteins.to_csv(dropped_path, index=False)

    # Return the filtered protein count data
    proteins_t = prot_with_sample_info.set_index("Sample ID")
    proteins_t = proteins_t.iloc[:, (sample_info.shape[1] - 1):]
    protein_count_data = proteins_t.T
    protein_count_data.columns.name = None
    protein_count_data.index.name = PRIMARY_KEY_COL

    protein_count_data.replace(0, np.nan, inplace=True)

    return protein_count_data

@debug
def remove_missing_proteins_global(protein_count_data: pd.DataFrame, threshold: float = 0.5) -> pd.DataFrame:
    max_nas = round(protein_count_data.shape[1] * threshold)

    #thresh argument is the max number of present values to drop the row
    post_drop = protein_count_data.dropna(thresh=(protein_count_data.shape[1] - max_nas), axis=0)

    dropped_proteins = pd.DataFrame(columns=[PRIMARY_KEY_COL])
    dropped = [prot for prot in protein_count_data.index if prot not in post_drop.index]
    for prot in dropped:
        dropped_proteins = pd.concat([
            dropped_proteins,
            pd.DataFrame({PRIMARY_KEY_COL: [prot]})
        ], ignore_index=True)

    # Save dropped proteins to CSV
    dropped_path = os.path.join(config["project_information"]["relative_path"], config["paths"]["dropped_proteins_list"])
    dropped_proteins.to_csv(dropped_path, index=False)
    
    return post_drop

@debug
def perform_simple_normalization(protein_count_data: pd.DataFrame, norm_func: Callable[[pd.Series], pd.Series]) -> pd.DataFrame:
    for sample in protein_count_data.columns:
        current_sample = protein_count_data[sample]
        sample_value = norm_func(current_sample)

        for protein in current_sample.index:
            protein_count_data.loc[protein, sample] -= sample_value

    return protein_count_data

@debug
def perform_total_value_normalization(protein_count_data: pd.DataFrame, ) -> pd.DataFrame:
    for sample in protein_count_data.columns:
        current_sample = protein_count_data[sample]
        total_value = current_sample.sum()

        for protein in current_sample.index:
            protein_count_data.loc[protein, sample] /= total_value

    return protein_count_data


@debug
def perform_quantile_normalization(protein_count_data: pd.DataFrame) -> pd.DataFrame:
    sorted_df = pd.DataFrame(np.sort(protein_count_data.values, axis=0), index=protein_count_data.index, columns=protein_count_data.columns)
    rank_means = sorted_df.mean(axis=1).values
    normalized_df = protein_count_data.rank(method="min").stack().astype(int).map(lambda rank: rank_means[rank - 1]).unstack()
    
    return normalized_df

@debug
def impute_half_value(protein_count_data: pd.DataFrame, axis: int, percentage_of_min: float) -> pd.DataFrame:
    # Iterate over each column or row based on the specified axis
    for i in (protein_count_data.columns if axis == 1 else protein_count_data.index):
        if axis == 1:  # Column-wise
            fill_value = protein_count_data[i].min() * percentage_of_min  # Min of the column times the percentage
            protein_count_data.loc[:, i] = protein_count_data.loc[:, i].fillna(fill_value)
        elif axis == 0:
            fill_value = protein_count_data.loc[i].min() * percentage_of_min  # Min of the row times the percentage
            protein_count_data.loc[i, :] = protein_count_data.loc[i, :].fillna(fill_value)
    
    return protein_count_data

@debug
def perform_log(protein_count_data: pd.DataFrame, log: int) -> pd.DataFrame:
    #Save data pre-transformation for certain analysis
    pre_trans_path = os.path.join(config['project_information']['relative_path'], config['paths']['pre_transformation'])
    protein_count_data.to_csv(pre_trans_path)

    #To avoid taking the log of any negative numbers, if negative numbers are present, we subtract the lowest negative number from all values
    #We subtract the negative to add
    if protein_count_data.min().min() < 0:
        protein_count_data -= (protein_count_data.min().min() - 1)

    #If there are any na's or 0s in the dataset, we replace the NaNs with 0 and add 1 (to avoid taking the log of 0)
    if (protein_count_data.iloc[:, 1:] == 0).any().any() or protein_count_data.iloc[:, 1:].isna().any().any():
        protein_count_data.replace(np.nan, 0, inplace=True)
        protein_count_data += 1

    match log:
        case 2:
            protein_count_data = np.log2(protein_count_data)

    return protein_count_data

def drop_duplicates(protein_count_data: pd.DataFrame) -> pd.DataFrame:
    return protein_count_data[~protein_count_data.index.duplicated(keep='first')].copy()

def get_worksheet_name(current_worksheet: int, n_worksheets: int):
    if current_worksheet == 0:
        return "Data"
    elif current_worksheet == 1:
        return "Sample Information"
    elif current_worksheet == (n_worksheets - 1):
        return "Removed Proteins"
    else:
        return f"Comparison {current_worksheet-1} Data"


@debug
def output_delivery_dataset(analysis_dataset: pd.DataFrame, protein_meta_data: pd.DataFrame):
    sample_info = load_intermediate_csv('sample_info')

    #split up analysis dataset columns -
    #   sheet 1: samples data + global missing + means
    #   sheet 2: sample groups
    #   sheet 3-n: comparison sheets
    #   sheet n+1: removed proteins w/ original counts

    sheets = []

    #sheet 1: samples data + global missing + means
    sheet_1 = analysis_dataset[[col for col in analysis_dataset if "Sample" in col]]

    sheet_1.reset_index(inplace=True)
    protein_meta_data[PRIMARY_KEY_COL] = protein_meta_data[PRIMARY_KEY_COL].astype(str)
    sheet_1[PRIMARY_KEY_COL] = sheet_1[PRIMARY_KEY_COL].astype(str)
    sheet_1 = protein_meta_data.merge(sheet_1, on=PRIMARY_KEY_COL, how="right")

    sheet_1["Assay Mean"] = sheet_1.iloc[:, len(protein_meta_data.columns):].apply(pd.to_numeric, errors='coerce').mean(axis=1)
    sheet_1["Global Missing"] = sheet_1.iloc[:, len(protein_meta_data.columns)-1:-1].isnull().sum(axis=1).astype(str) + " / " + str(sheet_1.shape[1] - 1)
    sheets.append(sheet_1)

    #sheet 2: sample groups
    sheets.append(sample_info)

    #sheet 3-n: comparison sheets
    comparison_num = 1
    while any([col.startswith(f"Comparison {comparison_num}:") for col in analysis_dataset.columns]):
        #first, find which samples are related to the comparision
        group_filtered_datasets = prepare_group_filtered_dict("post_pipeline", config['comparisons'][str(comparison_num)])

        samples_in_comparison_total = []
        for values in group_filtered_datasets.values():
            #I want to check which samples are in each group, and filter the eventual heatmap by those samples
            if values.get("Dataset") is not None:
                samples_in_dataset = [col for col in values["Dataset"].columns if col.startswith("Sample")]
                for sample in samples_in_dataset:
                    samples_in_comparison_total.append(sample)

        #we only want columns that are samples in the comparison or information about the comparison
        temp_sheet_cols = samples_in_comparison_total + [col for col in analysis_dataset.columns if col.startswith(f"Comparison {comparison_num}:")]
        sheets.append(analysis_dataset.loc[:, sorted(temp_sheet_cols, key=sort_key)])

        comparison_num += 1

    #sheet n+1: removed proteins w/ original counts
    raw_counts = load_intermediate_csv("no_processing")
    dropped_proteins = load_intermediate_csv("dropped_proteins_list")

    sheets.append(raw_counts.loc[dropped_proteins.index, :])

    #write to excel
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") #Grey background
    max_col_width = 90

    final_analysis_excel_path = os.path.join(DIR_PATH, "delivery/data/Data_With_Analysis.xlsx")
    with pd.ExcelWriter(final_analysis_excel_path, engine='openpyxl') as writer:
        workbook = writer.book

        n_metadata_col = len(protein_meta_data.columns) + 1

        #Data Sheet 1
        sheets[0].to_excel(writer, sheet_name=get_worksheet_name(0, len(sheets)), index=False, freeze_panes=SHEET1_FREEZE_PANES)
        worksheet = workbook[get_worksheet_name(0, len(sheets))]
        #For all columns
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = min((max_length + 2), max_col_width)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        #All cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        #For data cells (non-metadata)
        for row in worksheet.iter_rows(min_col=n_metadata_col, min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0000'

        # Apply the grey background to the top row
        for cell in worksheet[1]:  # The top row is row 1
            cell.fill = grey_fill
        
        #Data Sheet 2
        sheets[1].to_excel(writer, sheet_name=get_worksheet_name(1, len(sheets)), index=False, freeze_panes=(1,1))
        worksheet = workbook[get_worksheet_name(1, len(sheets))]

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = min((max_length + 2), max_col_width)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Apply the grey background to the top row
        for cell in worksheet[1]:  # The top row is row 1
            cell.fill = grey_fill


        #Comparison Sheets
        for i in range(2, len(sheets)-1):
            sheets[i].to_excel(writer, sheet_name=get_worksheet_name(i, len(sheets)), freeze_panes = (1,1))
            worksheet = workbook[get_worksheet_name(i, len(sheets))]

            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = min((max_length + 2), max_col_width)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            #All cells
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '0.0000'
                    else:
                        cell.alignment = Alignment(horizontal="center")

            # Apply the grey background to the top row
            for cell in worksheet[1]:  # The top row is row 1
                cell.fill = grey_fill

        #Removed proteins sheet
        sheets[-1].to_excel(writer, sheet_name=get_worksheet_name(len(sheets)-1, len(sheets)), freeze_panes=(1,1))
        worksheet = workbook[get_worksheet_name(len(sheets)-1, len(sheets))]

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = min((max_length + 2), max_col_width)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        #All cells
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '0.0000'
                else:
                    cell.alignment = Alignment(horizontal="center")

        # Apply the grey background to the top row
        for cell in worksheet[1]:  # The top row is row 1
            cell.fill = grey_fill